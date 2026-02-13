import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from git import Repo, InvalidGitRepositoryError, NoSuchPathError, GitCommandError
import datetime
import re
import subprocess
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
#################PROYECTO##################
__proyect__    = 'Git Branch Info & Recovery (with SQLite History)'
__author__     = "Mario Rubio"
__copyright__  = "Copyright 2022, The MRubioDev Project"
__credits__    = ["Mario Rubio"]
__license__    = "GPL"
__version__    = "V26.02.014"
__maintainer__ = "Mario Rubio"
__email__      = "https://mrubiodev.com/"
__status__     = "Development" #"Prototype", "Development", or "Production"


class GitBranchInfoApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"{__proyect__} - v{__version__} by {__author__}")
        self.root.geometry("1400x900")

        self.db_path = "git_branches.db" # Nombre de la base de datos SQLite

        self.root.grid_rowconfigure(0, weight=0)
        self.root.grid_rowconfigure(1, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        # --- Frame superior para la selección de carpeta ---
        top_frame = tk.Frame(root, padx=10, pady=10)
        top_frame.grid(row=0, column=0, sticky="ew")
        top_frame.grid_columnconfigure(1, weight=1)

        self.label_path = tk.Label(top_frame, text="Ruta del Repositorio Git:")
        self.label_path.grid(row=0, column=0, padx=5, pady=5, sticky="w")

        self.entry_path = tk.Entry(top_frame, width=70)
        self.entry_path.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        self.button_browse = tk.Button(top_frame, text="Examinar", command=self.browse_directory)
        self.button_browse.grid(row=0, column=2, padx=5, pady=5, sticky="e")

        self.button_get_info = tk.Button(top_frame, text="Obtener Información y Registrar", command=self.get_all_branch_info)
        self.button_get_info.grid(row=1, column=0, columnspan=3, pady=10)

        # --- Notebook (pestañas) ---
        self.notebook = ttk.Notebook(root)
        self.notebook.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)

        # Pestaña 1: Consola (salida de texto tradicional)
        self.console_frame = tk.Frame(self.notebook)
        self.notebook.add(self.console_frame, text="Consola")
        
        self.console_frame.grid_rowconfigure(0, weight=1)
        self.console_frame.grid_columnconfigure(0, weight=1)

        self.text_output = tk.Text(self.console_frame, wrap="word", state="disabled", font=("Consolas", 10))
        self.text_output.grid(row=0, column=0, sticky="nsew")

        console_scrollbar = tk.Scrollbar(self.console_frame, command=self.text_output.yview)
        console_scrollbar.grid(row=0, column=1, sticky="ns")
        self.text_output.config(yscrollcommand=console_scrollbar.set)

        # Pestaña 2: Búsqueda y Resultados en Tabla
        self.search_frame = tk.Frame(self.notebook)
        self.notebook.add(self.search_frame, text="Búsqueda en Base de Datos")
        
        self.search_frame.grid_rowconfigure(1, weight=1)
        self.search_frame.grid_columnconfigure(0, weight=1)

        # Frame de búsqueda
        search_controls = tk.Frame(self.search_frame, relief=tk.GROOVE, borderwidth=2, padx=10, pady=10)
        search_controls.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
        search_controls.grid_columnconfigure(1, weight=1)

        tk.Label(search_controls, text="Buscar por:", font=("Arial", 10, "bold")).grid(row=0, column=0, columnspan=4, pady=(0, 10), sticky="w")

        # Búsqueda por rama
        tk.Label(search_controls, text="Rama:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.entry_search_branch = tk.Entry(search_controls, width=30)
        self.entry_search_branch.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        
        # Búsqueda por ruta
        tk.Label(search_controls, text="Repositorio:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.entry_search_path = tk.Entry(search_controls, width=30)
        self.entry_search_path.grid(row=2, column=1, padx=5, pady=5, sticky="ew")
        
        # Búsqueda por archivo
        tk.Label(search_controls, text="Archivo:").grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.entry_search_file = tk.Entry(search_controls, width=30)
        self.entry_search_file.grid(row=3, column=1, padx=5, pady=5, sticky="ew")

        # Botones de búsqueda
        button_frame = tk.Frame(search_controls)
        button_frame.grid(row=4, column=0, columnspan=2, pady=10)
        
        self.button_search = tk.Button(button_frame, text="Buscar", command=self.perform_search, width=12)
        self.button_search.pack(side=tk.LEFT, padx=5)
        
        self.button_clear_search = tk.Button(button_frame, text="Limpiar Filtros", command=self.clear_search, width=12)
        self.button_clear_search.pack(side=tk.LEFT, padx=5)
        
        self.button_view_all = tk.Button(button_frame, text="Ver Todos", command=self.view_all_records, width=12)
        self.button_view_all.pack(side=tk.LEFT, padx=5)
        
        self.button_export = tk.Button(button_frame, text="Exportar Excel", command=self.export_to_excel, width=12)
        self.button_export.pack(side=tk.LEFT, padx=5)
        
        self.button_copy = tk.Button(button_frame, text="Copiar Selección", command=self.copy_selection, width=12)
        self.button_copy.pack(side=tk.LEFT, padx=5)

        # Tabla de resultados (Treeview)
        table_frame = tk.Frame(self.search_frame)
        table_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # Definir columnas
        columns = ("id", "repo", "branch", "type", "hash", "date", "author", "message", "files", "status")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", selectmode="extended")
        
        # Configurar encabezados
        self.tree.heading("id", text="ID")
        self.tree.heading("repo", text="Repositorio")
        self.tree.heading("branch", text="Rama")
        self.tree.heading("type", text="Tipo")
        self.tree.heading("hash", text="Hash")
        self.tree.heading("date", text="Fecha Commit")
        self.tree.heading("author", text="Autor")
        self.tree.heading("message", text="Mensaje")
        self.tree.heading("files", text="Archivos Modificados")
        self.tree.heading("status", text="Estado")

        # Configurar anchos de columna
        self.tree.column("id", width=40, anchor="center")
        self.tree.column("repo", width=150)
        self.tree.column("branch", width=150)
        self.tree.column("type", width=120)
        self.tree.column("hash", width=80)
        self.tree.column("date", width=130)
        self.tree.column("author", width=100)
        self.tree.column("message", width=200)
        self.tree.column("files", width=250)
        self.tree.column("status", width=80, anchor="center")

        self.tree.grid(row=0, column=0, sticky="nsew")

        # Scrollbars para la tabla
        tree_scrollbar_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        tree_scrollbar_y.grid(row=0, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=tree_scrollbar_y.set)

        tree_scrollbar_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        tree_scrollbar_x.grid(row=1, column=0, sticky="ew")
        self.tree.configure(xscrollcommand=tree_scrollbar_x.set)

        # Label para mostrar cantidad de resultados
        self.label_results = tk.Label(self.search_frame, text="Resultados: 0", font=("Arial", 9))
        self.label_results.grid(row=2, column=0, sticky="w", padx=10, pady=5)

        # Doble clic para ver detalles
        self.tree.bind("<Double-1>", self.show_details)
        
        # Menú contextual para la tabla principal
        self.tree_context_menu = tk.Menu(self.tree, tearoff=0)
        self.tree_context_menu.add_command(label="Copiar fila", command=self.copy_selection)
        self.tree_context_menu.add_command(label="Copiar todas las filas visibles", command=self.copy_all_visible)
        self.tree_context_menu.add_separator()
        self.tree_context_menu.add_command(label="Ver detalles", command=lambda: self.show_details(None))
        self.tree.bind("<Button-3>", self.show_context_menu)

        # Pestaña 3: Búsqueda en Lote
        self.batch_frame = tk.Frame(self.notebook)
        self.notebook.add(self.batch_frame, text="Búsqueda en Lote")
        
        self.batch_frame.grid_rowconfigure(2, weight=1)
        self.batch_frame.grid_columnconfigure(0, weight=1)

        # Instrucciones
        instructions = tk.Label(self.batch_frame, 
            text="Pega múltiples nombres de ramas o archivos (uno por línea o separados por comas)",
            font=("Arial", 10, "bold"), pady=10)
        instructions.grid(row=0, column=0, sticky="ew", padx=10)

        # Frame para tipo de búsqueda
        type_frame = tk.Frame(self.batch_frame)
        type_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=5)
        
        tk.Label(type_frame, text="Buscar por:").pack(side=tk.LEFT, padx=5)
        self.batch_search_type = tk.StringVar(value="branch")
        tk.Radiobutton(type_frame, text="Nombres de Ramas", variable=self.batch_search_type, 
                      value="branch").pack(side=tk.LEFT, padx=10)
        tk.Radiobutton(type_frame, text="Archivos Modificados", variable=self.batch_search_type, 
                      value="file").pack(side=tk.LEFT, padx=10)

        # Área de texto para entrada
        input_frame = tk.Frame(self.batch_frame)
        input_frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=5)
        input_frame.grid_rowconfigure(0, weight=1)
        input_frame.grid_columnconfigure(0, weight=1)
        
        tk.Label(input_frame, text="Lista de búsqueda:", anchor="w").grid(row=0, column=0, sticky="w", pady=(0, 5))
        
        self.batch_text_input = tk.Text(input_frame, height=10, wrap="word", font=("Consolas", 10))
        self.batch_text_input.grid(row=1, column=0, sticky="nsew")
        
        batch_input_scroll = tk.Scrollbar(input_frame, command=self.batch_text_input.yview)
        batch_input_scroll.grid(row=1, column=1, sticky="ns")
        self.batch_text_input.config(yscrollcommand=batch_input_scroll.set)

        # Botones
        button_batch_frame = tk.Frame(self.batch_frame)
        button_batch_frame.grid(row=3, column=0, pady=10)
        
        self.button_batch_search = tk.Button(button_batch_frame, text="Buscar en Lote", 
                                             command=self.perform_batch_search, width=15, height=2)
        self.button_batch_search.pack(side=tk.LEFT, padx=5)
        
        self.button_batch_clear = tk.Button(button_batch_frame, text="Limpiar", 
                                           command=self.clear_batch_search, width=12, height=2)
        self.button_batch_clear.pack(side=tk.LEFT, padx=5)
        
        self.button_batch_export = tk.Button(button_batch_frame, text="Exportar Excel", 
                                            command=self.export_batch_to_excel, width=12, height=2)
        self.button_batch_export.pack(side=tk.LEFT, padx=5)
        
        self.button_batch_copy = tk.Button(button_batch_frame, text="Copiar Resultados", 
                                          command=self.copy_batch_selection, width=12, height=2)
        self.button_batch_copy.pack(side=tk.LEFT, padx=5)

        # Área de resultados para búsqueda en lote
        results_frame = tk.Frame(self.batch_frame)
        results_frame.grid(row=4, column=0, sticky="nsew", padx=10, pady=5)
        results_frame.grid_rowconfigure(0, weight=1)
        results_frame.grid_columnconfigure(0, weight=1)
        
        tk.Label(results_frame, text="Resultados:", anchor="w", font=("Arial", 10, "bold")).grid(row=0, column=0, sticky="w", pady=(0, 5))
        
        # Tabla de resultados para lote
        batch_columns = ("item", "found", "repo", "branch", "type", "hash", "date", "files")
        self.batch_tree = ttk.Treeview(results_frame, columns=batch_columns, show="headings", selectmode="extended")
        
        self.batch_tree.heading("item", text="Búsqueda")
        self.batch_tree.heading("found", text="Encontrado")
        self.batch_tree.heading("repo", text="Repositorio")
        self.batch_tree.heading("branch", text="Rama")
        self.batch_tree.heading("type", text="Tipo")
        self.batch_tree.heading("hash", text="Hash")
        self.batch_tree.heading("date", text="Fecha")
        self.batch_tree.heading("files", text="Archivos")
        
        self.batch_tree.column("item", width=200)
        self.batch_tree.column("found", width=80, anchor="center")
        self.batch_tree.column("repo", width=120)
        self.batch_tree.column("branch", width=150)
        self.batch_tree.column("type", width=100)
        self.batch_tree.column("hash", width=80)
        self.batch_tree.column("date", width=130)
        self.batch_tree.column("files", width=250)
        
        self.batch_tree.grid(row=1, column=0, sticky="nsew")
        
        batch_scroll_y = ttk.Scrollbar(results_frame, orient="vertical", command=self.batch_tree.yview)
        batch_scroll_y.grid(row=1, column=1, sticky="ns")
        self.batch_tree.configure(yscrollcommand=batch_scroll_y.set)
        
        batch_scroll_x = ttk.Scrollbar(results_frame, orient="horizontal", command=self.batch_tree.xview)
        batch_scroll_x.grid(row=2, column=0, sticky="ew")
        self.batch_tree.configure(xscrollcommand=batch_scroll_x.set)
        
        self.label_batch_results = tk.Label(results_frame, text="Esperando búsqueda...", font=("Arial", 9))
        self.label_batch_results.grid(row=3, column=0, sticky="w", pady=5)
        
        # Doble clic en batch para ver detalles
        self.batch_tree.bind("<Double-1>", self.show_batch_details)
        
        # Menú contextual para tabla de lote
        self.batch_context_menu = tk.Menu(self.batch_tree, tearoff=0)
        self.batch_context_menu.add_command(label="Copiar fila", command=self.copy_batch_selection)
        self.batch_context_menu.add_command(label="Copiar todos los resultados", command=self.copy_all_batch)
        self.batch_context_menu.add_separator()
        self.batch_context_menu.add_command(label="Ver detalles", command=lambda: self.show_batch_details(None))
        self.batch_tree.bind("<Button-3>", self.show_batch_context_menu)

        # Inicializar la base de datos después de crear los componentes de la GUI
        self._init_database()

    def _init_database(self):
        """Inicializa la base de datos SQLite y crea las tablas si no existen."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Crear tabla si no existe
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS branches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                repo_path TEXT NOT NULL,
                branch_name TEXT NOT NULL,
                branch_type TEXT NOT NULL, -- 'remote_existing', 'local_existing', 'reflog_recoverable'
                last_commit_hash TEXT,
                commit_date TEXT,
                commit_message TEXT,
                commit_author TEXT,
                modified_files TEXT, -- Lista de archivos modificados separados por comas
                first_seen_date TEXT,
                last_updated_date TEXT,
                status TEXT, -- 'new', 'updated_commit', 'seen'
                UNIQUE(repo_path, branch_name, branch_type)
            )
        """)
        
        # Migración: Agregar columna modified_files si no existe
        cursor.execute("PRAGMA table_info(branches)")
        columns = [column[1] for column in cursor.fetchall()]
        
        if 'modified_files' not in columns:
            cursor.execute("ALTER TABLE branches ADD COLUMN modified_files TEXT DEFAULT ''")
            self.display_message("Base de datos actualizada: columna 'modified_files' agregada.", append=False)
        
        conn.commit()
        conn.close()
        self.display_message(f"Base de datos SQLite '{self.db_path}' inicializada.", append=False)

    def _record_branch_info(self, repo_path, branch_data):
        """Registra o actualiza la información de una rama en la base de datos."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        cursor.execute("""
            SELECT last_commit_hash FROM branches
            WHERE repo_path = ? AND branch_name = ? AND branch_type = ?
        """, (repo_path, branch_data['name'], branch_data['type']))
        
        existing_record = cursor.fetchone()
        
        status = "seen"
        if existing_record:
            if existing_record[0] != branch_data['hash']:
                status = "updated_commit"
                cursor.execute("""
                    UPDATE branches
                    SET last_commit_hash = ?, commit_date = ?, commit_message = ?, commit_author = ?, 
                        modified_files = ?, last_updated_date = ?, status = ?
                    WHERE repo_path = ? AND branch_name = ? AND branch_type = ?
                """, (branch_data['hash'], branch_data['date'], branch_data['message'], branch_data['author'], 
                      branch_data.get('files', ''), now, status,
                      repo_path, branch_data['name'], branch_data['type']))
            else:
                # Commit no ha cambiado, solo actualizamos la fecha de última vista
                cursor.execute("""
                    UPDATE branches
                    SET last_updated_date = ?, status = ?
                    WHERE repo_path = ? AND branch_name = ? AND branch_type = ?
                """, (now, status, repo_path, branch_data['name'], branch_data['type']))
        else:
            status = "new"
            cursor.execute("""
                INSERT INTO branches (repo_path, branch_name, branch_type, last_commit_hash, commit_date, 
                                     commit_message, commit_author, modified_files, first_seen_date, last_updated_date, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (repo_path, branch_data['name'], branch_data['type'], branch_data['hash'], branch_data['date'], 
                  branch_data['message'], branch_data['author'], branch_data.get('files', ''), now, now, status))
        
        conn.commit()
        conn.close()
        return status # Devolvemos el estado para mostrarlo en la UI

    def browse_directory(self):
        directory = filedialog.askdirectory()
        if directory:
            self.entry_path.delete(0, tk.END)
            self.entry_path.insert(0, directory)

    def display_message(self, message, append=False):
        self.text_output.config(state="normal")
        if not append:
            self.text_output.delete(1.0, tk.END)
        self.text_output.insert(tk.END, message + "\n")
        self.text_output.see(tk.END)
        self.text_output.config(state="disabled")
        self.root.update_idletasks()

    def get_all_branch_info(self):
        repo_path = self.entry_path.get()
        if not repo_path:
            messagebox.showwarning("Advertencia", "Por favor, selecciona la ruta del repositorio Git.")
            return

        self.display_message("Obteniendo información de ramas y registrando en la base de datos...", append=False)

        try:
            repo = Repo(repo_path)
            
            # --- 1. Obtener información de ramas remotas existentes ---
            existing_remote_branches_data = self._get_existing_remote_branches(repo, repo_path)
            self._display_branch_data(existing_remote_branches_data, "Ramas Remotas Existentes")

            # --- 2. Buscar ramas potencialmente recuperables en el reflog ---
            self.display_message("\n\n--- Buscando Ramas Potencialmente Recuperables en el Reflog Local ---", append=True)
            recoverable_branches_data = self._find_recoverable_branches_from_reflog(repo, repo_path, existing_remote_branches_data)
            self._display_branch_data(recoverable_branches_data, "Ramas Potencialmente Recuperables (Reflog)")

            self.display_message("\nProceso completado. La información ha sido registrada en la base de datos.", append=True)

        except InvalidGitRepositoryError:
            messagebox.showerror("Error", "La carpeta seleccionada no es un repositorio Git válido.")
            self.display_message("Error: La carpeta seleccionada no es un repositorio Git válido.", append=False)
        except NoSuchPathError:
            messagebox.showerror("Error", "La ruta especificada no existe.")
            self.display_message("Error: La ruta especificada no existe.", append=False)
        except Exception as e:
            messagebox.showerror("Error Inesperado", f"Ocurrió un error: {e}")
            self.display_message(f"Error inesperado: {e}", append=False)

    def _get_existing_remote_branches(self, repo, repo_path):
        """Obtiene la información de las ramas remotas existentes y la registra."""
        self.display_message("Realizando 'git fetch --prune origin' para actualizar la información remota...", append=True)
        
        if not repo.remotes:
            self.display_message("Error: El repositorio no tiene remotos configurados.", append=True)
            return []

        for remote in repo.remotes:
            try:
                remote.fetch(prune=True, progress=FetchProgress(self.text_output))
                self.display_message(f"Fetch completado para remoto: {remote.name}", append=True)
            except GitCommandError as e:
                self.display_message(f"Advertencia: No se pudo hacer fetch para el remoto '{remote.name}': {e}", append=True)
            except Exception as e:
                self.display_message(f"Advertencia: Error inesperado durante fetch para '{remote.name}': {e}", append=True)

        branches_data = []
        for remote_branch in repo.remote().refs:
            if remote_branch.name.endswith('/HEAD'):
                continue

            branch_name = remote_branch.name.split('/', 1)[1]
            
            try:
                last_commit = remote_branch.commit
                commit_hash_full = last_commit.hexsha  # Hash completo
                commit_hash = last_commit.hexsha[:7]   # Hash corto para mostrar
                commit_message = last_commit.message.strip().split('\n')[0]
                commit_date = datetime.datetime.fromtimestamp(last_commit.committed_date).strftime('%Y-%m-%d %H:%M:%S')
                author = last_commit.author.name
                
                # Obtener archivos modificados en el commit usando el hash completo
                files_modified = []
                try:
                    commit_obj = repo.commit(commit_hash_full)
                    if commit_obj.parents:
                        parent = commit_obj.parents[0]
                        diffs = parent.diff(commit_obj)
                        files_modified = [diff.a_path if diff.a_path else diff.b_path for diff in diffs]
                    else:
                        # Primer commit, todos los archivos son nuevos
                        files_modified = [item.path for item in commit_obj.tree.traverse() if item.type == 'blob']
                    
                    self.display_message(f"  Rama remota '{branch_name}': {len(files_modified)} archivos modificados encontrados", append=True)
                except Exception as e:
                    self.display_message(f"Advertencia: No se pudieron obtener archivos modificados para '{branch_name}' (commit {commit_hash}): {e}", append=True)
                
                files_str = ", ".join(files_modified[:10])  # Limitar a 10 archivos
                if len(files_modified) > 10:
                    files_str += f" ... (+{len(files_modified) - 10} más)"

                branch_info = {
                    "type": "remote_existing",
                    "name": branch_name,
                    "hash": commit_hash,
                    "date": commit_date,
                    "message": commit_message,
                    "author": author,
                    "files": files_str
                }
                status = self._record_branch_info(repo_path, branch_info)
                branch_info['status'] = status # Añadir el estado para mostrarlo
                branches_data.append(branch_info)
            except Exception as e:
                self.display_message(f"Advertencia: No se pudo obtener información para la rama remota '{branch_name}': {e}", append=True)
        
        return branches_data

    def _find_recoverable_branches_from_reflog(self, repo, repo_path, existing_remote_branches_data):
        """
        Busca en el reflog local commits que puedan pertenecer a ramas borradas
        y las registra en la base de datos.
        """
        self.display_message("Analizando el reflog local para encontrar ramas borradas...", append=True)
        
        recoverable_branches_data = []
        
        existing_branch_names = set(b['name'] for b in existing_remote_branches_data)
        existing_branch_names.update(h.name for h in repo.heads)

        patterns = [
            re.compile(r"HEAD@\{\d+\}:\s*(?:checkout|branch|merge|rebase\s+\(finish\)|rebase\s+\(pick\)|rebase\s+\(start\)):.*(?:from|to)\s+([^\s]+)$"),
            re.compile(r"HEAD@\{\d+\}:\s*commit(?:\s+\(initial\))?:.*refs/heads/([^\s]+)$"),
            re.compile(r"HEAD@\{\d+\}:\s*merge\s+([^\s]+):.*"),
            re.compile(r"HEAD@\{\d+\}:\s*reset:.*to\s+([^\s]+)$")
        ]

        try:
            result = subprocess.run(
                ['git', 'reflog', '--all'],
                cwd=repo.working_dir,
                capture_output=True,
                text=True,
                check=True,
                encoding='utf-8'
            )
            reflog_output = result.stdout
        except subprocess.CalledProcessError as e:
            self.display_message(f"Error al ejecutar 'git reflog --all': {e.stderr}", append=True)
            return []
        except Exception as e:
            self.display_message(f"Error inesperado al obtener el reflog: {e}", append=True)
            return []

        # Usamos un set para evitar procesar la misma rama recuperable varias veces
        processed_recoverable_branches = {} # {branch_name: commit_hash}

        for line in reflog_output.splitlines():
            match_hash = re.match(r"^([0-9a-f]{7,40})\s+.*", line)
            if not match_hash:
                continue
            
            commit_hash = match_hash.group(1)

            for pattern in patterns:
                match_branch_name = pattern.search(line)
                if match_branch_name:
                    branch_candidate = match_branch_name.group(1).strip()
                    
                    branch_candidate = branch_candidate.replace("refs/heads/", "")
                    branch_candidate = branch_candidate.replace("refs/remotes/origin/", "")
                    branch_candidate = branch_candidate.replace("refs/remotes/", "")
                    
                    if branch_candidate and \
                       branch_candidate not in existing_branch_names and \
                       not branch_candidate.startswith("HEAD") and \
                       not re.match(r"^[0-9a-f]{7,40}$", branch_candidate) and \
                       not branch_candidate.startswith("origin/") and \
                       branch_candidate not in processed_recoverable_branches: # Solo procesar una vez por rama
                        
                        try:
                            # Intentar obtener info del commit
                            commit_obj = repo.commit(commit_hash)
                            commit_message = commit_obj.message.strip().split('\n')[0]
                            commit_date = datetime.datetime.fromtimestamp(commit_obj.committed_date).strftime('%Y-%m-%d %H:%M:%S')
                            author = commit_obj.author.name
                            
                            # Obtener archivos modificados usando el hash completo del commit
                            files_modified = []
                            try:
                                if commit_obj.parents:
                                    parent = commit_obj.parents[0]
                                    diffs = parent.diff(commit_obj)
                                    files_modified = [diff.a_path if diff.a_path else diff.b_path for diff in diffs]
                                else:
                                    files_modified = [item.path for item in commit_obj.tree.traverse() if item.type == 'blob']
                                
                                self.display_message(f"  Rama '{branch_candidate}': {len(files_modified)} archivos modificados encontrados", append=True)
                            except Exception as e:
                                self.display_message(f"  Advertencia: Error al obtener archivos de '{branch_candidate}': {e}", append=True)
                            
                            files_str = ", ".join(files_modified[:10])
                            if len(files_modified) > 10:
                                files_str += f" ... (+{len(files_modified) - 10} más)"

                            branch_info = {
                                "type": "reflog_recoverable",
                                "name": branch_candidate,
                                "hash": commit_hash,
                                "date": commit_date,
                                "message": commit_message,
                                "author": author,
                                "files": files_str
                            }
                            status = self._record_branch_info(repo_path, branch_info)
                            branch_info['status'] = status
                            recoverable_branches_data.append(branch_info)
                            processed_recoverable_branches[branch_candidate] = commit_hash # Marcar como procesada
                        except Exception as e:
                            # Si el commit no es accesible (e.g., ya fue garbage collected), lo ignoramos
                            self.display_message(f"Advertencia: No se pudo obtener información del commit '{commit_hash}' para la rama '{branch_candidate}': {e}", append=True)
                    break
        
        return recoverable_branches_data

    def _display_branch_data(self, branches_data, title):
        """Muestra la información de las ramas en la UI, incluyendo el estado de la DB."""
        self.display_message(f"\n--- {title} ---", append=True)
        if not branches_data:
            self.display_message(f"No se encontraron {title.lower()} o no se pudo obtener su información.", append=True)
            return

        branches_data.sort(key=lambda x: x['date'], reverse=True)

        output_lines = []
        output_lines.append(f"{'Rama':<40} {'Hash':<8} {'Fecha Último Commit':<20} {'Estado DB':<12} {'Mensaje del Commit':<40}")
        output_lines.append("-" * 130)

        for info in branches_data:
            status_text = ""
            if info.get('status') == 'new':
                status_text = "NUEVA"
            elif info.get('status') == 'updated_commit':
                status_text = "ACTUALIZADA"
            else:
                status_text = "VISTA"

            output_lines.append(f"{info['name']:<40} {info['hash']:<8} {info['date']:<20} {status_text:<12} {info['message'][:37] + '...' if len(info['message']) > 40 else info['message']:<40}")
        
        self.display_message("\n".join(output_lines), append=True)

        # Para ramas recuperables, añadir los comandos
        if title == "Ramas Potencialmente Recuperables (Reflog)":
            self.display_message("\nComandos para recuperar las ramas marcadas como 'NUEVA' o 'ACTUALIZADA':", append=True)
            self.display_message("=" * 100, append=True)
            for info in branches_data:
                if info.get('status') in ['new', 'updated_commit']:
                    self.display_message(f"Rama: {info['name']}", append=True)
                    self.display_message(f"  Último commit conocido: {info['hash']}", append=True)
                    self.display_message(f"  Comandos para recuperar:", append=True)
                    self.display_message(f"    git branch \"{info['name']}\" {info['hash']}", append=True)
                    self.display_message(f"    git checkout \"{info['name']}\"", append=True)
                    self.display_message(f"    git push -u origin \"{info['name']}\"  (Opcional: para subirla al remoto)", append=True)
                    self.display_message("-" * 50, append=True)
            self.display_message("=" * 100, append=True)
            self.display_message("\nNota: Revisa cada rama recuperada antes de subirla al remoto.", append=True)

    def perform_search(self):
        """Realiza la búsqueda combinada según los filtros ingresados."""
        branch_name = self.entry_search_branch.get().strip()
        repo_path = self.entry_search_path.get().strip()
        file_name = self.entry_search_file.get().strip()
        
        if not branch_name and not repo_path and not file_name:
            messagebox.showwarning("Advertencia", "Por favor, ingresa al menos un criterio de búsqueda.")
            return

        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Construir la consulta dinámicamente
            query = """
                SELECT id, repo_path, branch_name, branch_type, last_commit_hash, commit_date, 
                       commit_message, commit_author, modified_files, status
                FROM branches
                WHERE 1=1
            """
            params = []
            
            if branch_name:
                query += " AND branch_name LIKE ?"
                params.append(f"%{branch_name}%")
            
            if repo_path:
                query += " AND repo_path LIKE ?"
                params.append(f"%{repo_path}%")
            
            if file_name:
                query += " AND modified_files LIKE ?"
                params.append(f"%{file_name}%")
            
            query += " ORDER BY commit_date DESC"
            
            self.display_message(f"Ejecutando búsqueda con criterios: Rama='{branch_name}', Repo='{repo_path}', Archivo='{file_name}'", append=True)
            
            cursor.execute(query, params)
            results = cursor.fetchall()
            conn.close()
            
            self.display_message(f"Se encontraron {len(results)} resultados.", append=True)
            self._populate_table(results)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al realizar la búsqueda: {e}")
            self.display_message(f"Error en búsqueda: {e}", append=True)

    def view_all_records(self):
        """Muestra todos los registros de la base de datos en la tabla."""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT id, repo_path, branch_name, branch_type, last_commit_hash, commit_date, 
                       commit_message, commit_author, modified_files, status
                FROM branches
                ORDER BY commit_date DESC
            """)
            
            results = cursor.fetchall()
            conn.close()
            
            self.display_message(f"Cargando todos los registros: {len(results)} encontrados.", append=True)
            self._populate_table(results)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar registros: {e}")
            self.display_message(f"Error al cargar registros: {e}", append=True)

    def clear_search(self):
        """Limpia los campos de búsqueda."""
        self.entry_search_branch.delete(0, tk.END)
        self.entry_search_path.delete(0, tk.END)
        self.entry_search_file.delete(0, tk.END)
        
        # Limpiar la tabla
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.label_results.config(text="Resultados: 0")

    def _populate_table(self, results):
        """Llena la tabla con los resultados de la búsqueda."""
        # Limpiar tabla existente
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        if not results:
            self.label_results.config(text="Resultados: 0 - No se encontraron registros")
            messagebox.showinfo("Búsqueda", "No se encontraron registros con los criterios especificados.")
            return
        
        # Llenar la tabla
        for row in results:
            try:
                row_id, repo_path, branch_name, branch_type, commit_hash, commit_date, commit_message, commit_author, modified_files, status = row
                
                # Manejar valores None
                modified_files = modified_files if modified_files else ""
                commit_message = commit_message if commit_message else ""
                commit_author = commit_author if commit_author else ""
                
                # Traducir tipo
                type_text = {
                    "remote_existing": "Remota",
                    "local_existing": "Local",
                    "reflog_recoverable": "Recuperable"
                }.get(branch_type, branch_type)
                
                # Traducir estado
                status_text = {
                    "new": "NUEVA",
                    "updated_commit": "ACTUALIZADA",
                    "seen": "VISTA"
                }.get(status, status if status else "")
                
                # Truncar repo path para mejor visualización
                repo_short = repo_path.split('\\')[-1] if '\\' in repo_path else repo_path.split('/')[-1]
                
                # Truncar mensaje si es muy largo
                message_short = commit_message[:50] + "..." if len(commit_message) > 50 else commit_message
                
                # Truncar archivos si es muy largo
                files_short = modified_files[:60] + "..." if len(modified_files) > 60 else modified_files
                
                self.tree.insert("", "end", values=(
                    row_id, repo_short, branch_name, type_text, commit_hash,
                    commit_date, commit_author, message_short, files_short, status_text
                ))
            except Exception as e:
                self.display_message(f"Error al procesar fila: {e}", append=True)
                continue
        
        self.label_results.config(text=f"Resultados: {len(results)}")
        
        # Cambiar a la pestaña de búsqueda
        self.notebook.select(self.search_frame)

    def show_details(self, event):
        """Muestra los detalles completos del registro seleccionado."""
        selection = self.tree.selection()
        if not selection:
            return
        
        item = self.tree.item(selection[0])
        row_id = item['values'][0]
        
        # Obtener datos completos de la base de datos
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT repo_path, branch_name, branch_type, last_commit_hash, commit_date, 
                   commit_message, commit_author, modified_files, first_seen_date, 
                   last_updated_date, status
            FROM branches
            WHERE id = ?
        """, (row_id,))
        
        result = cursor.fetchone()
        conn.close()
        
        if not result:
            return
        
        repo_path, branch_name, branch_type, commit_hash, commit_date, commit_message, commit_author, modified_files, first_seen, last_updated, status = result
        
        # Crear ventana de detalles
        details_window = tk.Toplevel(self.root)
        details_window.title(f"Detalles: {branch_name}")
        details_window.geometry("800x600")
        
        # Texto con detalles
        text_details = tk.Text(details_window, wrap="word", font=("Consolas", 10))
        text_details.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        details_scrollbar = tk.Scrollbar(details_window, command=text_details.yview)
        details_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        text_details.config(yscrollcommand=details_scrollbar.set)
        
        # Traducir tipo y estado
        type_text = {
            "remote_existing": "Remota Existente",
            "local_existing": "Local Existente",
            "reflog_recoverable": "Recuperable (Reflog)"
        }.get(branch_type, branch_type)
        
        status_text = {
            "new": "NUEVA",
            "updated_commit": "ACTUALIZADA",
            "seen": "VISTA"
        }.get(status, status)
        
        # Formatear detalles
        details = f"""
{'='*80}
DETALLES DE LA RAMA
{'='*80}

Nombre de la Rama: {branch_name}
Tipo: {type_text}
Estado: {status_text}

{'─'*80}
INFORMACIÓN DEL COMMIT
{'─'*80}

Hash del Commit: {commit_hash}
Fecha del Commit: {commit_date}
Autor: {commit_author}
Mensaje: {commit_message}

{'─'*80}
REPOSITORIO
{'─'*80}

Ruta: {repo_path}

{'─'*80}
ARCHIVOS MODIFICADOS
{'─'*80}

{modified_files if modified_files else 'No hay información de archivos'}

{'─'*80}
HISTORIAL DE REGISTRO
{'─'*80}

Primera vez visto: {first_seen}
Última actualización: {last_updated}

{'='*80}
"""
        
        if branch_type == "reflog_recoverable":
            details += f"""
COMANDOS PARA RECUPERAR LA RAMA
{'='*80}

git branch "{branch_name}" {commit_hash}
git checkout "{branch_name}"
git push -u origin "{branch_name}"

Nota: Revisa la rama antes de subirla al remoto.
"""
        
        text_details.insert("1.0", details)
        text_details.config(state="disabled")

    def perform_batch_search(self):
        """Realiza búsqueda en lote según el tipo seleccionado."""
        search_type = self.batch_search_type.get()
        input_text = self.batch_text_input.get("1.0", tk.END).strip()
        
        if not input_text:
            messagebox.showwarning("Advertencia", "Por favor, ingresa términos de búsqueda.")
            return
        
        # Limpiar tabla de resultados
        for item in self.batch_tree.get_children():
            self.batch_tree.delete(item)
        
        # Parsear la entrada: puede ser por líneas o separado por comas
        search_terms = []
        lines = input_text.split('\n')
        for line in lines:
            line = line.strip()
            if line:
                # Si tiene comas, dividir por comas
                if ',' in line:
                    terms = [t.strip() for t in line.split(',') if t.strip()]
                    search_terms.extend(terms)
                else:
                    search_terms.append(line)
        
        # Eliminar duplicados manteniendo el orden
        seen = set()
        unique_terms = []
        for term in search_terms:
            if term not in seen:
                seen.add(term)
                unique_terms.append(term)
        
        if not unique_terms:
            messagebox.showwarning("Advertencia", "No se encontraron términos válidos de búsqueda.")
            return
        
        self.display_message(f"\nIniciando búsqueda en lote de {len(unique_terms)} términos...", append=True)
        self.label_batch_results.config(text=f"Buscando {len(unique_terms)} términos...")
        self.root.update_idletasks()
        
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            total_found = 0
            total_not_found = 0
            
            for term in unique_terms:
                if search_type == "branch":
                    # Buscar por nombre de rama
                    cursor.execute("""
                        SELECT id, repo_path, branch_name, branch_type, last_commit_hash, 
                               commit_date, modified_files
                        FROM branches
                        WHERE branch_name LIKE ?
                        ORDER BY commit_date DESC
                    """, (f"%{term}%",))
                else:  # search_type == "file"
                    # Buscar por archivo modificado
                    cursor.execute("""
                        SELECT id, repo_path, branch_name, branch_type, last_commit_hash, 
                               commit_date, modified_files
                        FROM branches
                        WHERE modified_files LIKE ?
                        ORDER BY commit_date DESC
                    """, (f"%{term}%",))
                
                results = cursor.fetchall()
                
                if results:
                    total_found += 1
                    for row in results:
                        row_id, repo_path, branch_name, branch_type, commit_hash, commit_date, modified_files = row
                        
                        # Traducir tipo
                        type_text = {
                            "remote_existing": "Remota",
                            "local_existing": "Local",
                            "reflog_recoverable": "Recuperable"
                        }.get(branch_type, branch_type)
                        
                        # Truncar valores
                        repo_short = repo_path.split('\\')[-1] if '\\' in repo_path else repo_path.split('/')[-1]
                        files_short = (modified_files[:40] + "...") if modified_files and len(modified_files) > 40 else (modified_files or "")
                        
                        # Agregar a la tabla con tag para colorear
                        item_id = self.batch_tree.insert("", "end", values=(
                            term, "✓ SÍ", repo_short, branch_name, type_text, 
                            commit_hash, commit_date, files_short
                        ), tags=("found",))
                else:
                    total_not_found += 1
                    # Agregar como no encontrado
                    self.batch_tree.insert("", "end", values=(
                        term, "✗ NO", "-", "-", "-", "-", "-", "-"
                    ), tags=("not_found",))
                
                self.root.update_idletasks()
            
            conn.close()
            
            # Configurar colores para las tags
            self.batch_tree.tag_configure("found", background="#d4edda")  # Verde claro
            self.batch_tree.tag_configure("not_found", background="#f8d7da")  # Rojo claro
            
            self.label_batch_results.config(
                text=f"Búsqueda completada: {total_found} encontrados, {total_not_found} no encontrados (Total: {len(unique_terms)})"
            )
            self.display_message(f"Búsqueda en lote completada: {total_found}/{len(unique_terms)} términos encontrados", append=True)
            
            # Cambiar a la pestaña de búsqueda en lote
            self.notebook.select(self.batch_frame)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error en búsqueda en lote: {e}")
            self.display_message(f"Error en búsqueda en lote: {e}", append=True)

    def clear_batch_search(self):
        """Limpia la búsqueda en lote."""
        self.batch_text_input.delete("1.0", tk.END)
        for item in self.batch_tree.get_children():
            self.batch_tree.delete(item)
        self.label_batch_results.config(text="Esperando búsqueda...")

    def show_batch_details(self, event):
        """Muestra detalles del elemento seleccionado en búsqueda por lote."""
        selection = self.batch_tree.selection()
        if not selection:
            return
        
        item = self.batch_tree.item(selection[0])
        values = item['values']
        
        # Si no se encontró, no hay detalles que mostrar
        if values[1] == "✗ NO":
            messagebox.showinfo("No encontrado", f"El término '{values[0]}' no se encontró en la base de datos.")
            return
        
        # Buscar en la base de datos principal
        branch_name = values[3]
        
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT repo_path, branch_name, branch_type, last_commit_hash, commit_date, 
                       commit_message, commit_author, modified_files, first_seen_date, 
                       last_updated_date, status
                FROM branches
                WHERE branch_name = ?
                ORDER BY commit_date DESC
                LIMIT 1
            """, (branch_name,))
            
            result = cursor.fetchone()
            conn.close()
            
            if not result:
                return
            
            repo_path, branch_name, branch_type, commit_hash, commit_date, commit_message, commit_author, modified_files, first_seen, last_updated, status = result
            
            # Crear ventana de detalles (reutilizar la misma lógica)
            details_window = tk.Toplevel(self.root)
            details_window.title(f"Detalles: {branch_name}")
            details_window.geometry("800x600")
            
            text_details = tk.Text(details_window, wrap="word", font=("Consolas", 10))
            text_details.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
            
            details_scrollbar = tk.Scrollbar(details_window, command=text_details.yview)
            details_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            text_details.config(yscrollcommand=details_scrollbar.set)
            
            type_text = {
                "remote_existing": "Remota Existente",
                "local_existing": "Local Existente",
                "reflog_recoverable": "Recuperable (Reflog)"
            }.get(branch_type, branch_type)
            
            status_text = {
                "new": "NUEVA",
                "updated_commit": "ACTUALIZADA",
                "seen": "VISTA"
            }.get(status, status if status else "")
            
            details = f"""
{'='*80}
DETALLES DE LA RAMA
{'='*80}

Nombre de la Rama: {branch_name}
Tipo: {type_text}
Estado: {status_text}

{'─'*80}
INFORMACIÓN DEL COMMIT
{'─'*80}

Hash del Commit: {commit_hash}
Fecha del Commit: {commit_date}
Autor: {commit_author}
Mensaje: {commit_message if commit_message else 'N/A'}

{'─'*80}
REPOSITORIO
{'─'*80}

Ruta: {repo_path}

{'─'*80}
ARCHIVOS MODIFICADOS
{'─'*80}

{modified_files if modified_files else 'No hay información de archivos'}

{'─'*80}
HISTORIAL DE REGISTRO
{'─'*80}

Primera vez visto: {first_seen}
Última actualización: {last_updated}

{'='*80}
"""
            
            if branch_type == "reflog_recoverable":
                details += f"""
COMANDOS PARA RECUPERAR LA RAMA
{'='*80}

git branch "{branch_name}" {commit_hash}
git checkout "{branch_name}"
git push -u origin "{branch_name}"

Nota: Revisa la rama antes de subirla al remoto.
"""
            
            text_details.insert("1.0", details)
            text_details.config(state="disabled")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al obtener detalles: {e}")

    def show_context_menu(self, event):
        """Muestra el menú contextual para la tabla principal."""
        try:
            self.tree_context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.tree_context_menu.grab_release()

    def show_batch_context_menu(self, event):
        """Muestra el menú contextual para la tabla de lote."""
        try:
            self.batch_context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.batch_context_menu.grab_release()

    def copy_selection(self):
        """Copia las filas seleccionadas de la tabla principal al portapapeles."""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Por favor, selecciona al menos una fila.")
            return
        
        # Obtener encabezados
        headers = ["ID", "Repositorio", "Rama", "Tipo", "Hash", "Fecha Commit", 
                   "Autor", "Mensaje", "Archivos Modificados", "Estado"]
        
        # Construir texto con tabulaciones
        lines = ["\t".join(headers)]
        
        for item in selection:
            values = self.tree.item(item)['values']
            lines.append("\t".join(str(v) for v in values))
        
        text = "\n".join(lines)
        
        # Copiar al portapapeles
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.root.update()
        
        messagebox.showinfo("Copiado", f"{len(selection)} fila(s) copiada(s) al portapapeles.")

    def copy_all_visible(self):
        """Copia todas las filas visibles de la tabla principal al portapapeles."""
        items = self.tree.get_children()
        if not items:
            messagebox.showwarning("Advertencia", "No hay datos para copiar.")
            return
        
        # Obtener encabezados
        headers = ["ID", "Repositorio", "Rama", "Tipo", "Hash", "Fecha Commit", 
                   "Autor", "Mensaje", "Archivos Modificados", "Estado"]
        
        lines = ["\t".join(headers)]
        
        for item in items:
            values = self.tree.item(item)['values']
            lines.append("\t".join(str(v) for v in values))
        
        text = "\n".join(lines)
        
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.root.update()
        
        messagebox.showinfo("Copiado", f"{len(items)} fila(s) copiada(s) al portapapeles.")

    def copy_batch_selection(self):
        """Copia las filas seleccionadas de la tabla de lote al portapapeles."""
        selection = self.batch_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Por favor, selecciona al menos una fila.")
            return
        
        headers = ["Búsqueda", "Encontrado", "Repositorio", "Rama", "Tipo", "Hash", "Fecha", "Archivos"]
        lines = ["\t".join(headers)]
        
        for item in selection:
            values = self.batch_tree.item(item)['values']
            lines.append("\t".join(str(v) for v in values))
        
        text = "\n".join(lines)
        
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.root.update()
        
        messagebox.showinfo("Copiado", f"{len(selection)} fila(s) copiada(s) al portapapeles.")

    def copy_all_batch(self):
        """Copia todas las filas de la tabla de lote al portapapeles."""
        items = self.batch_tree.get_children()
        if not items:
            messagebox.showwarning("Advertencia", "No hay datos para copiar.")
            return
        
        headers = ["Búsqueda", "Encontrado", "Repositorio", "Rama", "Tipo", "Hash", "Fecha", "Archivos"]
        lines = ["\t".join(headers)]
        
        for item in items:
            values = self.batch_tree.item(item)['values']
            lines.append("\t".join(str(v) for v in values))
        
        text = "\n".join(lines)
        
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.root.update()
        
        messagebox.showinfo("Copiado", f"{len(items)} fila(s) copiada(s) al portapapeles.")

    def export_to_excel(self):
        """Exporta los datos de la tabla principal a un archivo Excel."""
        items = self.tree.get_children()
        if not items:
            messagebox.showwarning("Advertencia", "No hay datos para exportar.")
            return
        
        # Solicitar ubicación de guardado
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Guardar como Excel"
        )
        
        if not file_path:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Búsqueda de Ramas"
            
            # Encabezados
            headers = ["ID", "Repositorio", "Rama", "Tipo", "Hash", "Fecha Commit", 
                       "Autor", "Mensaje", "Archivos Modificados", "Estado"]
            
            # Estilo para encabezados
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Datos
            for row_num, item in enumerate(items, 2):
                values = self.tree.item(item)['values']
                for col_num, value in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = str(value)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 8   # ID
            ws.column_dimensions['B'].width = 25  # Repositorio
            ws.column_dimensions['C'].width = 30  # Rama
            ws.column_dimensions['D'].width = 15  # Tipo
            ws.column_dimensions['E'].width = 12  # Hash
            ws.column_dimensions['F'].width = 18  # Fecha
            ws.column_dimensions['G'].width = 20  # Autor
            ws.column_dimensions['H'].width = 40  # Mensaje
            ws.column_dimensions['I'].width = 50  # Archivos
            ws.column_dimensions['J'].width = 12  # Estado
            
            wb.save(file_path)
            messagebox.showinfo("Éxito", f"Datos exportados exitosamente a:\n{file_path}")
            self.display_message(f"Datos exportados a: {file_path}", append=True)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar a Excel: {e}")
            self.display_message(f"Error al exportar: {e}", append=True)

    def export_batch_to_excel(self):
        """Exporta los datos de la tabla de búsqueda en lote a Excel."""
        items = self.batch_tree.get_children()
        if not items:
            messagebox.showwarning("Advertencia", "No hay datos para exportar.")
            return
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Guardar como Excel"
        )
        
        if not file_path:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Búsqueda en Lote"
            
            headers = ["Búsqueda", "Encontrado", "Repositorio", "Rama", "Tipo", "Hash", "Fecha", "Archivos"]
            
            # Estilo para encabezados
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Datos con colores según si fue encontrado
            found_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            not_found_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            
            for row_num, item in enumerate(items, 2):
                values = self.batch_tree.item(item)['values']
                is_found = values[1] == "✓ SÍ"
                
                for col_num, value in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = str(value)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    
                    # Aplicar color de fondo
                    if is_found:
                        cell.fill = found_fill
                    else:
                        cell.fill = not_found_fill
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 35  # Búsqueda
            ws.column_dimensions['B'].width = 12  # Encontrado
            ws.column_dimensions['C'].width = 25  # Repositorio
            ws.column_dimensions['D'].width = 30  # Rama
            ws.column_dimensions['E'].width = 15  # Tipo
            ws.column_dimensions['F'].width = 12  # Hash
            ws.column_dimensions['G'].width = 18  # Fecha
            ws.column_dimensions['H'].width = 50  # Archivos
            
            wb.save(file_path)
            messagebox.showinfo("Éxito", f"Datos exportados exitosamente a:\n{file_path}")
            self.display_message(f"Datos de búsqueda en lote exportados a: {file_path}", append=True)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar a Excel: {e}")
            self.display_message(f"Error al exportar: {e}", append=True)

    def search_by_branch(self):
        """Método obsoleto - mantener por compatibilidad."""
        self.perform_search()

    def search_by_path(self):
        """Método obsoleto - mantener por compatibilidad."""
        self.perform_search()

# Clase para mostrar el progreso del fetch en el Text widget
class FetchProgress(object):
    def __init__(self, text_widget):
        self.text_widget = text_widget
        self.current_line_start = None

    def __call__(self, op_code, cur_count, max_count=None, message=''):
        if max_count:
            progress = f" ({cur_count}/{max_count})"
        else:
            progress = f" ({cur_count})"

        full_message = f"  {op_code} {message}{progress}"
        
        self.text_widget.config(state="normal")
        if self.current_line_start:
            # Si ya hay una línea de progreso, la borramos para sobrescribir
            self.text_widget.delete(self.current_line_start, tk.END)
        else:
            # Si es la primera vez, añadimos un salto de línea si no es el inicio
            if self.text_widget.index(tk.END) != "1.0":
                self.text_widget.insert(tk.END, "\n")
            self.current_line_start = self.text_widget.index(tk.END)

        self.text_widget.insert(tk.END, full_message)
        self.text_widget.see(tk.END)
        self.text_widget.config(state="disabled")
        self.text_widget.update_idletasks()

        # Si la operación ha terminado (cur_count == max_count), reseteamos para la próxima línea
        if max_count and cur_count == max_count:
            self.current_line_start = None

if __name__ == "__main__":
    root = tk.Tk()
    app = GitBranchInfoApp(root)
    root.mainloop()